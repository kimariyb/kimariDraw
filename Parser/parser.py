import numpy as np
import openpyxl
import pandas as pd


class Parser:
    def __init__(self, filename):
        self.filename = filename

    def parse(self):
        raise NotImplementedError


class TXTParser(Parser):
    def parse(self):
        # 从 txt 文件中读取数据
        data = np.loadtxt(self.filename)
        # 将数据转换为 DataFrame 对象
        df = pd.DataFrame(data, columns=['x', 'y'])
        return df

    def parse_multi(self):
        # 从 txt 文件中读取数据
        data = np.loadtxt(self.filename)
        # 将数据转换为 DataFrame 对象
        df = pd.DataFrame(data, columns=['x'] + ['y{}'.format(i) for i in range(1, data.shape[1])])

        return df


class ExcelParser(Parser):
    def parse(self):
        # 打开 Excel 文件
        wb = openpyxl.load_workbook(self.filename)
        # 选择工作表
        ws = wb.active
        # 读取数据
        data = []
        for row in ws.iter_rows(values_only=True):
            if all(v is not None for v in row):
                data.append({'x': float(row[0]), 'y': float(row[1])})

        # 将数据转换为 DataFrame 对象
        df = pd.DataFrame(data)
        return df

    def parse_multi(self):
        # 打开 Excel 文件
        wb = openpyxl.load_workbook(self.filename)
        # 选择工作表
        ws = wb.active
        # 读取数据
        data = []
        for row in ws.iter_rows(values_only=True):
            if all(v is not None for v in row):
                data.append([float(row[0]), *[float(v) for v in row[1:]]])

        # 将数据转换为 DataFrame 对象
        df = pd.DataFrame(data, columns=['x'] + ['y{}'.format(i) for i in range(1, len(data[0]))])

        return df
